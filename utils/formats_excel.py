import requests

import xlsxwriter

import numpy as np
import pandas as pd
import PyPDF2
from PyPDF2 import PdfFileReader

import base64
import io
import datetime
from styleframe import StyleFrame

from UliPlot.XLSX import auto_adjust_xlsx_column_width
from utils.transform import Transform #  get worksheet
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill

from openpyxl.styles.differential import DifferentialStyle



import xlsxwriter
import openpyxl

class Formats():

    def __init__(self, dataframe, xlsFilepath):
        
        self.transf = Transform(dataframe, xlsFilepath)
        print(self.transf)
        
        
        
        
    def formats_xlsxwriter(self, workbook = xlsxwriter.workbook.Workbook()): # Add some cell formats.
        
        #workbook  = writer.book
        #worksheet = writer.sheets['Summary__IA_BTP']
                
        # Light red fill with dark red text.
        format1 = workbook.add_format({'bg_color':   '#e67583',
                                       'font_color': '#9C0006'})

        format2 = workbook.add_format({'num_format': '0%'})

        format_basic = workbook.add_format({'num_format': '#,##0.00 €', 'font_color': 'black'})

        format3 = workbook.add_format({'num_format': '0%',
                                      'bold': False,
                                      'border': 1,
                                      'fg_color': '#ffcccc',
                                      'font_color': 'black'  })



        format5 = workbook.add_format({'num_format': '0%',
                                      'bold': True,
                                      'border': 1,
                                      'fg_color': '#ffb266',
                                      'font_color': 'black'  })



        # Add a header format.
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D7E4BC',
            'border': 1})

        # Add a number format for cells with money.
        money_fmt_low = workbook.add_format({'num_format': '#,##0.00 €', 'border': 1, 'bold' : False,
                                        'bg_color': '#8ae3be', 'font_color': 'black'
                                        })

        # Add a number format for cells with money.
        money_fmt_high = workbook.add_format({'num_format': '#,##0.00 €', 'border': 1, 'bold' : True,
                                        'bg_color': '#66b2ff', 'font_color': 'black'
                                        })

        # Add a number format for cells with money.
        money_fmt_bet = workbook.add_format({'num_format': '#,##0.00 €', 'border': 1, 'bold' : True,
                                        'bg_color': '#ff752b', 'font_color': 'black'
                                        })

        dic = {}
        dic["format1"] = format1
        dic["format2"] = format2
        dic["format_basic"] = format_basic
        dic["format3"] = format3
        dic["format5"] = format5
        dic["header_format"] = header_format
        dic["money_fmt_low"] = money_fmt_low
        dic["money_fmt_high"] = money_fmt_high
        dic["money_fmt_bet"] = money_fmt_bet

        return dic
    
    
    
    
    def appply_formats_xlsxwriter(self, worksheet, workbook = xlsxwriter.workbook.Workbook()): # Add some cell formats.
            
            dict_formats = self.formats_xlsxwriter(workbook = workbook)
            
            df = self.transf.dataframe
            
            for col in list(range(1,2)) + list(range(4,6)) + [7]:
                worksheet.set_column(col, col, 50)


            for col in [2]: # n°facture
                worksheet.set_column(col, col, 15, dict_formats["format3"])


            for col in [3]: # Montant
                worksheet.set_column(col, col, 15,  dict_formats["format_basic"]) #, money_fmt)


            for col in [6]: # n°sem
                worksheet.set_column(col, col, 7)


            # Write the column headers with the defined format.
            for col_num, value in enumerate(df.columns[0:2].values):
                worksheet.write(0, col_num, value, dict_formats["header_format"])


            first_col = 0
            last_col = 0
            last_row = df.shape[0]

            worksheet.set_column(first_col, first_col, 25)

            worksheet.conditional_format('A2:A10', {'type': '3_color_scale'})

            worksheet.conditional_format('G2:G1093', {'type':      '3_color_scale',
                                         'min_value': 5,
                                         'mid_value': 15,
                                         'max_value': 32,
                                         'min_color': '#CC0000',
                                         'mid_color': '#0994eb',
                                         'max_color': '#09eb7e'})

            # This is the correct syntax.
            worksheet.conditional_format('E2:F9' ,
                {'type':     'formula',
                 'criteria': '=OR($B2<$C2,AND($B2="",$C2>TODAY()))',
                 'format':   dict_formats["format1"]})


            worksheet.conditional_format('D2:D1093', {'type':     'cell',
                                                'criteria': '>',
                                                'value':    200_000,
                                                'format':   dict_formats["money_fmt_high"]})

            worksheet.conditional_format('D2:D1093', {'type':     'cell',
                                                'criteria': '<',
                                                'value':    2500,
                                                'format':   dict_formats["money_fmt_low"]})

            worksheet.conditional_format('D2:D1093', {'type':     'cell',
                                            'criteria': 'between',
                                            'minimum':  5000,
                                            'maximum':  7000,
                                            'format':   dict_formats["money_fmt_bet"]})
            
            return (workbook, worksheet)


            
    def formats_openpyxl(self) : #, workbook = openpyxl.workbook.workbook.Workbook()): # Add some cell formats.
        
        
        format1 = ColorScaleRule(start_type="min",
                                   start_color="00FF0000",  # Red
                                   end_type="max",
                                   end_color="0000FF00")  # Green




        format2 = ColorScaleRule(start_type="num",
                                   start_value=1,
                                   start_color="00FF0000",  # Red
                                   mid_type="num",
                                   mid_value=3,
                                   mid_color="00FFFF00",  # Yellow
                                   end_type="num",
                                   end_value=5,
                                   end_color="0000FF00")  # Green

        # Create a few styles
        bold_font = Font(bold=True)
        big_red_text = Font(color="00FF0000", size=10)
        center_aligned_text = Alignment(horizontal="center")
        double_border_side = Side(border_style="double")
        square_border = Border(top=double_border_side,
                            right=double_border_side,
                            bottom=double_border_side,
                            left=double_border_side)
        
        
        red_background = PatternFill(fgColor="00FF0000")
        diff_style = DifferentialStyle(fill=red_background)
        rule = Rule(type="expression", dxf=diff_style)
        
        pinkFill = PatternFill(start_color='ffcccc', #FFFF0000',
                   end_color='ffcccc', #FFFF0000',
                   fill_type='solid')
        
        
        orangeFill = PatternFill(start_color='ff9966', #FFFF0000',
                   end_color='ff9966', #FFFF0000',
                   fill_type='solid')
        
        greenFill = PatternFill(start_color='d7e4bc', #FFFF0000',
                   end_color='d7e4bc', #FFFF0000',
                   fill_type='solid')
        
        blueFill = PatternFill(start_color='66b2ff', #FFFF0000',
                   end_color='66b2ff', #FFFF0000',
                   fill_type='solid')
        
        

        dic = {}
        dic["format1"] = format1
        dic["format2"] = format2
        dic["bold_font"] = bold_font
        dic["big_red_text"] = big_red_text
        dic["center_aligned_text"] = center_aligned_text
        dic["double_border_side"] = double_border_side
        dic["square_border"] = square_border
        dic["red_background"] = red_background
        dic["diff_style"] = diff_style
        dic["rule"] = rule
        dic["pinkFill"] = pinkFill
        dic["orangeFill"] = orangeFill
        dic["greenFill"] = greenFill
        dic["blueFill"] = blueFill

        return dic
        
    def appply_formats_openpyxl(self, worksheet, sheetname="Summary__IA_BTP"):#, workbook = openpyxl.workbook.Workbook()): # Add some cell formats.
            
        dict_formats = self.formats_openpyxl()#workbook = workbook)
        
        print("worksheet   ", worksheet)

        # Again, let's add this gradient to the star ratings, column "H"
        worksheet[sheetname].conditional_formatting.add("A2:A100", dict_formats["format1"])
        worksheet[sheetname].conditional_formatting.add("C2:C100", dict_formats["format2"])
        
        
        worksheet[sheetname]['A1'].fill = dict_formats["greenFill"]
        worksheet[sheetname]['C1'].fill = dict_formats["greenFill"]
        
        for col in worksheet[sheetname].iter_cols(min_row=3, min_col = 5, max_col=6, max_row=5):
            for cell in col:
                cell.fill = dict_formats["pinkFill"]
                
                
        for col in worksheet[sheetname].iter_cols(min_row=1, min_col = 7, max_col=7, max_row=1093):
            for cell in col:
                cell.fill = dict_formats["orangeFill"]
        
        for col in worksheet[sheetname].iter_cols(min_row=2, min_col = 4, max_col=4, max_row=1093):
            for cell in col:
                cell.fill = dict_formats["blueFill"]
                
                
                
        # set the height of the row
        worksheet[sheetname].row_dimensions[1].height = 50

        # set the width of the column
        
        for col in ['A', 'B', 'E', 'F', 'H']:
            worksheet[sheetname].column_dimensions[col].width = 20
        
        for col in ['C', 'D']:
            worksheet[sheetname].column_dimensions[col].width = 15
                   
        
        return worksheet # worksheet['Summary__IA_BTP']#worksheet

        
